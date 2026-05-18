import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl

from phast_temperature_analyser.core.types import TemperatureType


class ExcelProcessor:
    """Handles Excel file processing and data extraction."""
    
    def __init__(self, temperature_type: TemperatureType, verbose: bool = False):
        self.temperature_type = temperature_type
        self.verbose = verbose
        self.logger = logging.getLogger(__name__)
    
    def process_files(self, folder_path: str) -> List[Dict[str, Any]]:
        """Process all Excel files in the given folder."""
        excel_files = list(Path(folder_path).glob("*.xlsx"))
        all_data = []
        
        for file_path in excel_files:
            try:
                file_data = self._process_single_file(file_path)
                all_data.extend(file_data)
            except Exception as e:
                self.logger.error(f"Error processing {file_path}: {e}")
                
        return all_data
    
    def _process_single_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process a single Excel file."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        file_data = []
        
        if self.verbose:
            self.logger.info(f"Processing file: {file_path}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = self._analyze_sheet(ws)
            file_data.extend(sheet_data)
            
        wb.close()
        return file_data
    
    def _analyze_sheet(self, ws) -> List[Dict[str, Any]]:
        """Analyze a single worksheet for dispersion data."""
        sheet_data = []
        equipment_item = None
        scenario = None
        weather = None
        
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                value = str(cell.value).strip() if cell.value else ""
                
                if value.startswith("Equipment Item:"):
                    equipment_item = value.split(":", 1)[1].strip()
                elif value.startswith("Scenario ("):
                    scenario = value.split(":", 1)[1].strip()
                elif value.startswith("Weather:"):
                    weather = value.split(":", 1)[1].strip()
                elif value == "Time-varying Observer Dispersion Data (before along-wind-diffusion effects)":
                    data = self._extract_dispersion_data(ws, cell.row)
                    if data and equipment_item and scenario and weather:
                        sheet_data.append({
                            'equipment_item': equipment_item,
                            'scenario': scenario,
                            'weather': weather,
                            'distances': data['distances'],
                            'temperatures': data['temperatures']
                        })
        
        return sheet_data
    
    def _extract_dispersion_data(self, ws, start_row: int) -> Optional[Dict[str, List]]:
        """Extract dispersion data from the worksheet."""
        try:
            # Get headers (2 rows after the title)
            headers = [cell.value for cell in ws[start_row + 2]]
            
            # Find column indices
            distance_col = headers.index("Downwind distance [m]")
            temp_header = (
                "C/Line vapour temperature [degC]" if self.temperature_type == TemperatureType.VAPOUR
                else "C/Line liquid temperature [degC]"
            )
            temp_col = headers.index(temp_header)
            
            # Extract data
            distances, temperatures = [], []
            
            for row_num in range(start_row + 3, ws.max_row + 1):
                row_data = [cell.value for cell in ws[row_num]]
                
                if not row_data or row_data[0] != 1:
                    break
                
                distance = row_data[distance_col]
                temperature = row_data[temp_col]
                
                if distance is not None and temperature is not None:
                    try:
                        distances.append(float(distance))
                        temperatures.append(float(temperature))
                    except (ValueError, TypeError):
                        continue
            
            return {
                'distances': distances,
                'temperatures': temperatures
            } if distances and temperatures else None
            
        except (ValueError, IndexError) as e:
            self.logger.error(f"Error extracting dispersion data: {e}")
            return None 